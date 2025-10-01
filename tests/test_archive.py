import os
import zipfile
from io import TextIOWrapper
from pypdf import PdfReader
from openpyxl import load_workbook
import pandas as pd


def test_archive(create_archive):
    assert os.path.exists(create_archive)
    assert zipfile.is_zipfile(create_archive)


def test_archive_contains_tmp(create_archive):
    with zipfile.ZipFile(create_archive, 'r') as zip_file:
        file_list = zip_file.namelist()

        assert 'file_example_pdf.pdf' in file_list
        assert 'file_example_XLSX.xlsx' in file_list
        assert 'file_example_csv.csv' in file_list


def test_pdf(create_archive):
    with zipfile.ZipFile(create_archive, 'r') as zip_file:
        with zip_file.open('file_example_pdf.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            first_page = reader.pages[1]
            text = first_page.extract_text()

            assert "Simple, Rapid, Effective, and Scalable" in text


def test_xlsx(create_archive):
    with zipfile.ZipFile(create_archive, 'r') as zip_file:
        with zip_file.open('file_example_XLSX.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.cell(row=1, column=2).value == "First Name"


def test_csv(create_archive):
    with zipfile.ZipFile(create_archive, 'r') as zip_file:
        with zip_file.open('file_example_csv.csv') as csv_file:
            reader_csv = pd.read_csv(TextIOWrapper(csv_file, encoding='utf-8'))
            column_names = list(reader_csv.columns)
            assert ['Пример файла'] == column_names